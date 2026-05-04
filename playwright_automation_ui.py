"""
Playwright Automation UI — PySide6
Upload Excel → Run Automation → Write Results → Export Updated Excel
Styled with TUIQ-inspired light theme (white panels, teal/blue ribbon, panel headers)
"""

import sys
import os
import asyncio
import threading
from datetime import datetime
from pathlib import Path

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QTextEdit, QProgressBar, QFrame, QSplitter, QHeaderView,
    QStatusBar, QMessageBox, QComboBox, QLineEdit, QGroupBox,
    QGridLayout, QScrollArea, QSizePolicy, QToolButton,
    QCheckBox
)
from PySide6.QtCore import (
    Qt, QThread, Signal, QObject, QSize, QRect
)
from PySide6.QtGui import (
    QColor, QPalette, QFont, QIcon, QPixmap, QPainter,
    QPen, QAction
)

# ── Optional dependencies (graceful fallback) ──────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from playwright.async_api import async_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False

# ── Colour palette  (TUIQ light theme) ────────────────────────────────────
BG          = "#F0F2F5"
SURFACE     = "#FFFFFF"
PANEL_HDR   = "#E8ECF0"
BORDER      = "#CDD1D6"
BORDER_LT   = "#E2E5E9"

TEAL        = "#00897B"
TEAL_DARK   = "#00695C"
TEAL_LIGHT  = "#E0F2F1"
TEAL_BORDER = "#80CBC4"

BLUE        = "#0078D4"
BLUE_DARK   = "#005A9E"
BLUE_LIGHT  = "#EBF4FC"

GREEN       = "#2E7D32"
GREEN_LIGHT = "#E8F5E9"

RED         = "#C62828"
RED_LIGHT   = "#FFEBEE"

AMBER       = "#F57F17"
AMBER_LIGHT = "#FFFDE7"

TEXT        = "#1A1A2E"
TEXT_SEC    = "#4A5568"
TEXT_MUTED  = "#8A9BB0"

STATUS_BAR  = "#0078D4"

FONT_UI   = "'Segoe UI', 'SF Pro Text', 'Helvetica Neue', sans-serif"
FONT_MONO = "'Cascadia Code', 'JetBrains Mono', 'Consolas', monospace"


# ══════════════════════════════════════════════════════════════════════════════
#  Sample Excel generator
# ══════════════════════════════════════════════════════════════════════════════

def create_sample_excel(path: str) -> None:
    if not HAS_OPENPYXL:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Automation Tasks"
    header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    header_font = XLFont(bold=True, color="FFFFFF", size=11)
    headers = ["ID", "URL", "Action", "Selector", "Input Value", "Status", "Result", "Timestamp"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    sample_rows = [
        [1, "https://example.com",         "navigate",    "",     "", "Pending", "", ""],
        [2, "https://example.com",         "screenshot",  "",     "", "Pending", "", ""],
        [3, "https://quotes.toscrape.com", "scrape_title","h1",   "", "Pending", "", ""],
        [4, "https://httpbin.org/get",     "get_content", "body", "", "Pending", "", ""],
        [5, "https://example.com",         "get_links",   "a",    "", "Pending", "", ""],
    ]
    for row_data in sample_rows:
        ws.append(row_data)
    for col, width in zip("ABCDEFGH", [5, 35, 15, 20, 20, 10, 30, 20]):
        ws.column_dimensions[chr(64 + col)].width = width
    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════════
#  Worker — runs in a QThread
# ══════════════════════════════════════════════════════════════════════════════

class AutomationWorker(QObject):
    log        = Signal(str, str)
    row_update = Signal(int, str, str)
    progress   = Signal(int)
    finished   = Signal(bool, str)

    def __init__(self, tasks: list[dict], headless: bool = True):
        super().__init__()
        self.tasks    = tasks
        self.headless = headless
        self._stop    = False

    def stop(self):
        self._stop = True

    def run(self):
        asyncio.run(self._run_async())

    async def _run_async(self):
        total = len(self.tasks)
        if total == 0:
            self.finished.emit(False, "No tasks to run.")
            return

        if not HAS_PLAYWRIGHT:
            self.log.emit("Playwright not installed — running in simulation mode.", "warn")
            for i, task in enumerate(self.tasks):
                if self._stop:
                    break
                await asyncio.sleep(0.6)
                status = "Done"
                result = f"[Simulated] {task.get('action','?')} on {task.get('url','?')}"
                self.row_update.emit(i, status, result)
                self.progress.emit(int((i + 1) / total * 100))
                self.log.emit(f"Row {i+1}: {result}", "info")
            self.finished.emit(True, "Simulation complete.")
            return

        self.log.emit("Launching browser…", "info")
        try:
            async with async_playwright() as pw:
                browser = await pw.chromium.launch(headless=self.headless)
                page    = await browser.new_page()
                for i, task in enumerate(self.tasks):
                    if self._stop:
                        self.log.emit("Stopped by user.", "warn")
                        break
                    url      = task.get("url", "")
                    action   = task.get("action", "").lower()
                    selector = task.get("selector", "")
                    value    = task.get("input_value", "")
                    self.log.emit(f"Row {i+1}: {action} → {url}", "info")
                    status = "Done"
                    result = ""
                    try:
                        if action == "navigate":
                            await page.goto(url, timeout=15000)
                            result = page.url
                        elif action == "screenshot":
                            await page.goto(url, timeout=15000)
                            fname = f"screenshot_{i+1}_{datetime.now():%H%M%S}.png"
                            await page.screenshot(path=fname)
                            result = f"Saved: {fname}"
                        elif action == "scrape_title":
                            await page.goto(url, timeout=15000)
                            if selector:
                                el = page.locator(selector).first
                                result = await el.inner_text()
                            else:
                                result = await page.title()
                        elif action == "get_content":
                            await page.goto(url, timeout=15000)
                            el = page.locator(selector or "body").first
                            result = (await el.inner_text())[:120].replace("\n", " ")
                        elif action == "get_links":
                            await page.goto(url, timeout=15000)
                            links = await page.eval_on_selector_all(
                                selector or "a", "els => els.map(e => e.href)"
                            )
                            result = f"{len(links)} links found"
                        elif action == "fill":
                            await page.goto(url, timeout=15000)
                            await page.fill(selector, value)
                            result = "Filled"
                        else:
                            result = f"Unknown action: {action}"
                            status = "Skipped"
                    except Exception as ex:
                        status = "Error"
                        result = str(ex)[:100]
                        self.log.emit(f"Error: {ex}", "error")
                    self.row_update.emit(i, status, result)
                    self.progress.emit(int((i + 1) / total * 100))
                    self.log.emit(f"{status}: {result}", "info" if status == "Done" else "error")
                await browser.close()
            self.finished.emit(True, "Automation complete.")
        except Exception as ex:
            self.finished.emit(False, str(ex))


# ══════════════════════════════════════════════════════════════════════════════
#  Reusable styled widgets
# ══════════════════════════════════════════════════════════════════════════════

class RibbonButton(QToolButton):
    def __init__(self, emoji: str, label: str, accent: str = BLUE, parent=None):
        super().__init__(parent)
        self._accent = accent
        self.setCursor(Qt.PointingHandCursor)
        self.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
        self.setText(label)
        self.setFixedSize(62, 58)
        pix = self._emoji_pixmap(emoji, 24)
        self.setIcon(QIcon(pix))
        self.setIconSize(QSize(24, 24))
        self.setStyleSheet(f"""
            QToolButton {{
                background: transparent;
                border: 1px solid transparent;
                border-radius: 4px;
                color: {TEXT};
                font-size: 10px;
                font-family: {FONT_UI};
                padding: 2px 1px 1px 1px;
            }}
            QToolButton:hover {{
                background: {accent}18;
                border-color: {accent}55;
                color: {accent};
            }}
            QToolButton:pressed  {{ background: {accent}30; }}
            QToolButton:disabled {{ color: {TEXT_MUTED}; }}
        """)

    @staticmethod
    def _emoji_pixmap(text: str, size: int) -> QPixmap:
        pix = QPixmap(size, size)
        pix.fill(Qt.transparent)
        p = QPainter(pix)
        p.setFont(QFont("Segoe UI Emoji", int(size * 0.65)))
        p.drawText(QRect(0, 0, size, size), Qt.AlignCenter, text)
        p.end()
        return pix


class RibbonSep(QFrame):
    def __init__(self):
        super().__init__()
        self.setFrameShape(QFrame.VLine)
        self.setFixedWidth(1)
        self.setFixedHeight(50)
        self.setStyleSheet(f"background: {BORDER}; border: none;")


class RibbonGroup(QWidget):
    def __init__(self, title: str, parent=None):
        super().__init__(parent)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(6, 3, 6, 0)
        lay.setSpacing(0)
        self.btn_row = QHBoxLayout()
        self.btn_row.setSpacing(2)
        lay.addLayout(self.btn_row, 1)
        lbl = QLabel(title)
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setStyleSheet(f"""
            font-size: 9px; color: {TEXT_MUTED}; font-family: {FONT_UI};
            border-top: 1px solid {BORDER}; padding-top: 2px; margin-top: 1px;
        """)
        lay.addWidget(lbl)

    def add(self, w: QWidget):
        self.btn_row.addWidget(w)


class PanelHeader(QWidget):
    def __init__(self, title: str, subtitle: str = "", colour: str = TEAL, parent=None):
        super().__init__(parent)
        self.setFixedHeight(58)
        self.setStyleSheet(f"background: {colour}; border: none;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(12, 8, 12, 6)
        lay.setSpacing(1)
        t = QLabel(title)
        t.setStyleSheet(f"font-size: 13px; font-weight: 700; color: #FFFFFF; font-family: {FONT_UI};")
        lay.addWidget(t)
        if subtitle:
            s = QLabel(subtitle)
            s.setStyleSheet(f"font-size: 10px; color: #FFFFFFCC; font-family: {FONT_UI};")
            s.setWordWrap(True)
            lay.addWidget(s)


class StyledButton(QPushButton):
    @staticmethod
    def _darken(hex_colour: str, factor: float = 0.82) -> str:
        h = hex_colour.lstrip("#")
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        return f"#{int(r*factor):02X}{int(g*factor):02X}{int(b*factor):02X}"

    def __init__(self, text: str, accent: str = TEAL, parent=None):
        super().__init__(text, parent)
        self._accent = accent
        self.setMinimumHeight(34)
        self.setCursor(Qt.PointingHandCursor)
        a_hover = self._darken(accent, 0.82)
        a_press = self._darken(accent, 0.65)
        self.setStyleSheet(f"""
            QPushButton {{
                background: {accent}; color: #FFFFFF;
                border: none; border-radius: 5px;
                padding: 0 18px; font-size: 12px; font-weight: 600;
                font-family: {FONT_UI};
            }}
            QPushButton:hover   {{ background: {a_hover}; color: #FFFFFF; }}
            QPushButton:pressed {{ background: {a_press}; color: #FFFFFF; }}
            QPushButton:disabled {{ background: {PANEL_HDR}; color: {TEXT_MUTED}; }}
        """)


class GhostButton(QPushButton):
    def __init__(self, text: str, parent=None):
        super().__init__(text, parent)
        self.setMinimumHeight(34)
        self.setCursor(Qt.PointingHandCursor)
        self.setStyleSheet(f"""
            QPushButton {{
                background: {SURFACE}; color: {TEXT_SEC};
                border: 1px solid {BORDER}; border-radius: 5px;
                padding: 0 14px; font-size: 12px; font-family: {FONT_UI};
            }}
            QPushButton:hover {{
                background: {BLUE_LIGHT}; border-color: {BLUE}; color: {BLUE};
            }}
            QPushButton:pressed  {{ background: {BORDER}; }}
            QPushButton:disabled {{ color: {TEXT_MUTED}; }}
        """)


class StatusBadge(QLabel):
    COLOURS = {
        "Pending": (AMBER,      AMBER_LIGHT),
        "Done":    (GREEN,      GREEN_LIGHT),
        "Error":   (RED,        RED_LIGHT),
        "Skipped": (TEXT_MUTED, PANEL_HDR),
        "Running": (BLUE,       BLUE_LIGHT),
    }

    def __init__(self, status: str = "Pending", parent=None):
        super().__init__(status, parent)
        self.set_status(status)

    def set_status(self, status: str):
        fg, bg = self.COLOURS.get(status, (TEXT_MUTED, PANEL_HDR))
        self.setText(status)
        self.setStyleSheet(f"""
            QLabel {{
                background: {bg}; color: {fg};
                border: 1px solid {fg}66; border-radius: 4px;
                padding: 1px 8px; font-size: 11px; font-weight: 600;
                font-family: {FONT_UI};
            }}
        """)
        self.setAlignment(Qt.AlignCenter)


# ══════════════════════════════════════════════════════════════════════════════
#  Main Window
# ══════════════════════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Playwright Automation Studio")
        self.resize(1280, 820)
        self.setMinimumSize(900, 600)

        self._excel_path: str | None = None
        self._tasks:      list[dict] = []
        self._worker:     AutomationWorker | None = None
        self._thread:     QThread | None = None
        self._results:    list[dict] = []

        self._apply_global_style()
        self._build_ui()
        self._log("Welcome to Playwright Automation Studio.", "info")
        self._log("1. Load an Excel file  →  2. Run automation  →  3. Export results.", "muted")

    # ── Global stylesheet ─────────────────────────────────────────────────────

    def _apply_global_style(self):
        self.setStyleSheet(f"""
            QMainWindow, QWidget {{
                background: {BG};
                color: {TEXT};
                font-family: {FONT_UI};
                font-size: 12px;
            }}
            QSplitter::handle {{ background: {BORDER}; }}
            QScrollBar:vertical {{
                background: {BG}; width: 8px; border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {BORDER}; border-radius: 4px; min-height: 20px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
            QScrollBar:horizontal {{ background: {BG}; height: 8px; border: none; }}
            QScrollBar::handle:horizontal {{ background: {BORDER}; border-radius: 4px; }}
            QHeaderView::section {{
                background: {PANEL_HDR}; color: {TEXT_SEC};
                border: none; border-bottom: 1px solid {BORDER};
                border-right: 1px solid {BORDER_LT};
                padding: 5px 10px; font-size: 11px; font-weight: 600;
                text-transform: uppercase; letter-spacing: 0.4px;
            }}
            QTableWidget {{
                background: {SURFACE}; border: 1px solid {BORDER};
                border-radius: 6px; gridline-color: {BORDER_LT};
                selection-background-color: {BLUE_LIGHT};
                selection-color: {BLUE_DARK};
            }}
            QTableWidget::item {{
                padding: 5px 10px; border-bottom: 1px solid {BORDER_LT}; color: {TEXT};
            }}
            QTableWidget::item:selected {{ background: {BLUE_LIGHT}; color: {BLUE_DARK}; }}
            QTableWidget::item:alternate {{ background: {BG}; }}
            QComboBox {{
                background: {SURFACE}; color: {TEXT};
                border: 1px solid {BORDER}; border-radius: 5px;
                padding: 5px 10px; min-height: 30px;
            }}
            QComboBox:focus {{ border-color: {BLUE}; }}
            QComboBox::drop-down {{ border: none; width: 18px; }}
            QComboBox QAbstractItemView {{
                background: {SURFACE}; color: {TEXT};
                selection-background-color: {BLUE_LIGHT};
                selection-color: {BLUE_DARK}; border: 1px solid {BORDER};
            }}
            QLineEdit {{
                background: {SURFACE}; color: {TEXT};
                border: 1px solid {BORDER}; border-radius: 5px;
                padding: 5px 10px; min-height: 30px;
            }}
            QLineEdit:focus {{ border-color: {BLUE}; }}
            QGroupBox {{
                background: {SURFACE}; border: 1px solid {BORDER};
                border-radius: 6px; margin-top: 10px; padding-top: 6px;
                font-weight: 600; font-size: 12px; color: {TEXT_SEC};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin; left: 12px;
                padding: 0 6px; background: {SURFACE};
            }}
            QProgressBar {{
                background: {PANEL_HDR}; border: none; border-radius: 3px; height: 5px;
            }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {TEAL}, stop:1 {BLUE});
                border-radius: 3px;
            }}
            QStatusBar {{
                background: {STATUS_BAR}; color: #FFFFFF;
                padding: 3px 12px; font-size: 11px;
            }}
            QMenuBar {{
                background: {SURFACE}; color: {TEXT};
                border-bottom: 1px solid {BORDER}; padding: 1px 6px;
            }}
            QMenuBar::item {{ padding: 4px 10px; border-radius: 3px; }}
            QMenuBar::item:selected {{ background: {BLUE_LIGHT}; color: {BLUE}; }}
            QMenuBar::item:pressed  {{ background: {BLUE}; color: white; }}
            QMenu {{
                background: {SURFACE}; border: 1px solid {BORDER};
                border-radius: 4px; padding: 4px;
            }}
            QMenu::item {{ padding: 6px 22px; border-radius: 3px; }}
            QMenu::item:selected {{ background: {BLUE_LIGHT}; color: {BLUE}; }}
            QCheckBox {{ spacing: 6px; color: {TEXT_SEC}; font-size: 12px; }}
            QCheckBox::indicator {{
                width: 14px; height: 14px; border: 1px solid {BORDER};
                border-radius: 3px; background: {SURFACE};
            }}
            QCheckBox::indicator:checked {{ background: {TEAL}; border-color: {TEAL_DARK}; }}
        """)

    # ── Build UI ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        self._build_menubar()

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._build_ribbon())

        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(1)
        splitter.addWidget(self._build_left_panel())
        splitter.addWidget(self._build_right_panel())
        splitter.setSizes([820, 440])
        root.addWidget(splitter, 1)

        root.addWidget(self._build_bottom_bar())
        self.statusBar().showMessage("  Ready")

    # ── Menu bar ──────────────────────────────────────────────────────────────

    def _build_menubar(self):
        mb = self.menuBar()
        file_m = mb.addMenu("File")
        file_m.addAction(QAction("Open Excel…",     self, triggered=self._browse_excel))
        file_m.addAction(QAction("Generate Sample", self, triggered=self._create_sample))
        file_m.addSeparator()
        file_m.addAction(QAction("Export Results",  self, triggered=self._export_excel))
        file_m.addSeparator()
        file_m.addAction(QAction("Exit",            self, triggered=self.close))
        mb.addMenu("Run")
        mb.addMenu("View")
        mb.addMenu("Settings")
        corner = QLabel("Playwright Automation Studio  v1.0")
        corner.setStyleSheet(f"font-size: 11px; color: {TEXT_MUTED}; padding-right: 12px;")
        mb.setCornerWidget(corner, Qt.TopRightCorner)

    # ── Ribbon ────────────────────────────────────────────────────────────────

    def _build_ribbon(self) -> QWidget:
        ribbon = QFrame()
        ribbon.setFixedHeight(84)
        ribbon.setStyleSheet(f"""
            QFrame {{
                background: {SURFACE};
                border-bottom: 2px solid {BORDER};
            }}
        """)
        lay = QHBoxLayout(ribbon)
        lay.setContentsMargins(8, 3, 8, 0)
        lay.setSpacing(0)

        # Excel group
        grp_excel = RibbonGroup("Excel")
        btn_open   = RibbonButton("📂", "Open\nExcel",  TEAL)
        btn_sample = RibbonButton("📋", "Sample\nFile", BLUE)
        btn_open.clicked.connect(self._browse_excel)
        btn_sample.clicked.connect(self._create_sample)
        grp_excel.add(btn_open)
        grp_excel.add(btn_sample)
        lay.addWidget(grp_excel)
        lay.addWidget(RibbonSep())

        # Browser group
        grp_browser = RibbonGroup("Browser")
        inner = QWidget()
        inner_lay = QVBoxLayout(inner)
        inner_lay.setContentsMargins(4, 6, 4, 4)
        inner_lay.setSpacing(5)
        self.browser_mode = QComboBox()
        self.browser_mode.addItems(["Headless (background)", "Headed (visible)"])
        self.browser_mode.setFixedWidth(165)
        inner_lay.addWidget(self.browser_mode)
        grp_browser.add(inner)
        lay.addWidget(grp_browser)
        lay.addWidget(RibbonSep())

        # Run group
        grp_run = RibbonGroup("Run Options")
        self._rib_run  = RibbonButton("▶",  "Execute", TEAL)
        self._rib_stop = RibbonButton("⛔", "Stop",    RED)
        self._rib_run.clicked.connect(self._run_automation)
        self._rib_stop.clicked.connect(self._stop_automation)
        self._rib_stop.setEnabled(False)
        grp_run.add(self._rib_run)
        grp_run.add(self._rib_stop)
        lay.addWidget(grp_run)
        lay.addWidget(RibbonSep())

        # Export group
        grp_export = RibbonGroup("Results")
        self._rib_export = RibbonButton("💾", "Export\nExcel", GREEN)
        self._rib_export.setEnabled(False)
        self._rib_export.clicked.connect(self._export_excel)
        btn_clr = RibbonButton("🗑", "Clear\nLog", AMBER)
        btn_clr.clicked.connect(lambda: self.log_box.clear())
        grp_export.add(self._rib_export)
        grp_export.add(btn_clr)
        lay.addWidget(grp_export)
        lay.addWidget(RibbonSep())

        # Library badges
        badges_w = QWidget()
        bl = QVBoxLayout(badges_w)
        bl.setContentsMargins(10, 10, 10, 10)
        bl.setSpacing(5)
        for lib, ok in [("Playwright", HAS_PLAYWRIGHT), ("openpyxl", HAS_OPENPYXL)]:
            badge = QLabel(f"  {'✔' if ok else '✖'}  {lib}  ")
            badge.setStyleSheet(f"""
                background: {GREEN_LIGHT if ok else RED_LIGHT};
                color: {GREEN if ok else RED};
                border: 1px solid {GREEN if ok else RED}55;
                border-radius: 4px; padding: 2px 6px;
                font-size: 11px; font-weight: 600;
            """)
            bl.addWidget(badge)
        lay.addWidget(badges_w)
        lay.addStretch()
        return ribbon

    # ── Left panel ────────────────────────────────────────────────────────────

    def _build_left_panel(self) -> QWidget:
        panel = QWidget()
        panel.setStyleSheet(f"background: {BG};")
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        lay.addWidget(PanelHeader(
            "Automation Tasks",
            "Load an Excel file, configure options, then run automation.",
            TEAL
        ))

        body = QWidget()
        body_lay = QVBoxLayout(body)
        body_lay.setContentsMargins(14, 14, 10, 14)
        body_lay.setSpacing(12)

        # File section
        file_group = QGroupBox("Excel File")
        fg_lay = QVBoxLayout(file_group)
        fg_lay.setSpacing(8)
        file_row = QHBoxLayout()
        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet(f"""
            background: {BG}; border: 1px dashed {BORDER};
            border-radius: 5px; padding: 7px 12px; color: {TEXT_MUTED};
        """)
        self.file_label.setMinimumHeight(34)
        btn_browse = GhostButton("Browse…")
        btn_browse.setFixedWidth(88)
        btn_browse.clicked.connect(self._browse_excel)
        btn_sample = GhostButton("Sample")
        btn_sample.setFixedWidth(78)
        btn_sample.clicked.connect(self._create_sample)
        btn_sample.setToolTip("Generate a sample Excel file")
        file_row.addWidget(self.file_label, 1)
        file_row.addWidget(btn_browse)
        file_row.addWidget(btn_sample)
        fg_lay.addLayout(file_row)
        body_lay.addWidget(file_group)

        # Options section
        opt_group = QGroupBox("Options")
        og_lay = QGridLayout(opt_group)
        og_lay.setSpacing(8)
        og_lay.setColumnMinimumWidth(0, 110)
        og_lay.addWidget(QLabel("Output path:"), 0, 0)
        self.output_path = QLineEdit()
        self.output_path.setPlaceholderText("Auto (same folder as input)")
        og_lay.addWidget(self.output_path, 0, 1)
        body_lay.addWidget(opt_group)

        # Tasks header + count badge
        tasks_hdr = QHBoxLayout()
        tbl_lbl = QLabel("Tasks")
        tbl_lbl.setStyleSheet(f"font-size: 12px; font-weight: 700; color: {TEAL_DARK};")
        self._task_count_lbl = QLabel("0 rows")
        self._task_count_lbl.setStyleSheet(f"""
            background: {TEAL_LIGHT}; color: {TEAL_DARK};
            border: 1px solid {TEAL_BORDER}; border-radius: 10px;
            padding: 1px 10px; font-size: 11px; font-weight: 600;
        """)
        tasks_hdr.addWidget(tbl_lbl)
        tasks_hdr.addSpacing(8)
        tasks_hdr.addWidget(self._task_count_lbl)
        tasks_hdr.addStretch()
        body_lay.addLayout(tasks_hdr)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["#", "URL", "Action", "Selector", "Status", "Result"])
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        self.table.setAlternatingRowColors(True)
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        hh.setSectionResizeMode(5, QHeaderView.Stretch)
        for c, w in [(0, 36), (2, 100), (3, 100), (4, 80)]:
            self.table.setColumnWidth(c, w)
        self.table.setMinimumHeight(200)
        body_lay.addWidget(self.table, 1)

        lay.addWidget(body, 1)
        return panel

    # ── Right panel ───────────────────────────────────────────────────────────

    def _build_right_panel(self) -> QWidget:
        panel = QWidget()
        panel.setStyleSheet(f"background: {BG};")
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        lay.addWidget(PanelHeader(
            "Console & Summary",
            "Live log output and run statistics.",
            "#37474F"
        ))

        body = QWidget()
        body_lay = QVBoxLayout(body)
        body_lay.setContentsMargins(10, 14, 14, 14)
        body_lay.setSpacing(12)

        log_hdr = QLabel("Console Output")
        log_hdr.setStyleSheet(f"font-size: 12px; font-weight: 700; color: {TEXT_SEC};")
        body_lay.addWidget(log_hdr)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet(f"""
            QTextEdit {{
                background: #1E2A35;
                border: 1px solid {BORDER};
                border-radius: 6px;
                padding: 10px;
                font-family: {FONT_MONO};
                font-size: 12px;
                color: #E6EDF3;
            }}
        """)
        body_lay.addWidget(self.log_box, 1)

        # Stats cards
        stats_group = QGroupBox("Run Summary")
        sg = QGridLayout(stats_group)
        sg.setSpacing(8)

        self.stat_total   = self._stat_label("0", TEXT_SEC)
        self.stat_done    = self._stat_label("0", GREEN)
        self.stat_error   = self._stat_label("0", RED)
        self.stat_pending = self._stat_label("0", AMBER)

        for col, (lbl_text, val_w, bg, accent) in enumerate([
            ("Total",   self.stat_total,   PANEL_HDR,   TEXT_SEC),
            ("Done",    self.stat_done,    GREEN_LIGHT, GREEN),
            ("Error",   self.stat_error,   RED_LIGHT,   RED),
            ("Pending", self.stat_pending, AMBER_LIGHT, AMBER),
        ]):
            card = QFrame()
            card.setStyleSheet(f"""
                QFrame {{
                    background: {bg};
                    border: 1px solid {accent}44;
                    border-top: 3px solid {accent};
                    border-radius: 6px;
                }}
            """)
            cl = QVBoxLayout(card)
            cl.setSpacing(2)
            cl.setContentsMargins(6, 6, 6, 6)
            lbl = QLabel(lbl_text)
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet(f"font-size: 10px; color: {TEXT_MUTED}; font-weight: 600; background: transparent; border: none;")
            cl.addWidget(lbl)
            val_w.setStyleSheet(val_w.styleSheet() + " background: transparent; border: none;")
            cl.addWidget(val_w)
            sg.addWidget(card, 0, col)

        body_lay.addWidget(stats_group)
        lay.addWidget(body, 1)
        return panel

    def _stat_label(self, text: str, colour: str = TEXT) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(f"font-size: 22px; font-weight: 700; color: {colour};")
        lbl.setAlignment(Qt.AlignCenter)
        return lbl

    # ── Bottom bar ────────────────────────────────────────────────────────────

    def _build_bottom_bar(self) -> QWidget:
        bar = QFrame()
        bar.setFixedHeight(54)
        bar.setStyleSheet(f"""
            QFrame {{
                background: {SURFACE};
                border-top: 1px solid {BORDER};
            }}
        """)
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(16, 0, 16, 0)
        lay.setSpacing(10)

        prog_col = QVBoxLayout()
        prog_col.setSpacing(3)
        prog_lbl = QLabel("Progress")
        prog_lbl.setStyleSheet(f"font-size: 10px; color: {TEXT_MUTED};")
        self.progress = QProgressBar()
        self.progress.setFixedHeight(5)
        self.progress.setTextVisible(False)
        self.progress.setValue(0)
        prog_col.addWidget(prog_lbl)
        prog_col.addWidget(self.progress)

        self.btn_run = StyledButton("▶  Run Automation", TEAL)
        self.btn_run.setFixedWidth(162)
        self.btn_run.clicked.connect(self._run_automation)

        self.btn_stop = StyledButton("⛔  Stop", RED)
        self.btn_stop.setFixedWidth(100)
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self._stop_automation)

        self.btn_export = StyledButton("⬇  Export Excel", GREEN)
        self.btn_export.setFixedWidth(140)
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self._export_excel)

        btn_clear = GhostButton("Clear Log")
        btn_clear.setFixedWidth(90)
        btn_clear.clicked.connect(self.log_box.clear)

        lay.addLayout(prog_col, 1)
        lay.addWidget(btn_clear)
        lay.addWidget(self.btn_stop)
        lay.addWidget(self.btn_run)
        lay.addWidget(self.btn_export)
        return bar

    # ── Excel I/O ─────────────────────────────────────────────────────────────

    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)"
        )
        if path:
            self._load_excel(path)

    def _create_sample(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Sample Excel", "sample_tasks.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "Missing Library",
                "openpyxl is not installed.\n\npip install openpyxl")
            return
        create_sample_excel(path)
        self._log(f"Sample Excel created: {path}", "info")
        self._load_excel(path)

    def _load_excel(self, path: str):
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "Missing Library",
                "openpyxl is not installed.\n\npip install openpyxl")
            return
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                self._log("Excel file is empty.", "warn")
                return

            headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
            self._tasks   = []
            self._results = []
            self.table.setRowCount(0)

            for ri, row in enumerate(rows[1:], 1):
                task = dict(zip(headers, row))
                task["_row_index"] = ri - 1
                self._tasks.append(task)
                self._results.append({"status": "Pending", "result": ""})

                r = self.table.rowCount()
                self.table.insertRow(r)
                self.table.setRowHeight(r, 34)
                self.table.setItem(r, 0, self._cell(str(ri)))
                self.table.setItem(r, 1, self._cell(str(task.get("url", ""))))
                self.table.setItem(r, 2, self._cell(str(task.get("action", ""))))
                self.table.setItem(r, 3, self._cell(str(task.get("selector", ""))))
                badge = StatusBadge("Pending")
                self.table.setCellWidget(r, 4, badge)
                self.table.setItem(r, 5, self._cell(""))

            self._excel_path = path
            fname = Path(path).name
            self.file_label.setText(f"📄  {fname}")
            self.file_label.setStyleSheet(f"""
                background: {TEAL_LIGHT}; border: 1px solid {TEAL_BORDER};
                border-radius: 5px; padding: 7px 12px;
                color: {TEAL_DARK}; font-weight: 600;
            """)
            n = len(self._tasks)
            self._task_count_lbl.setText(f"{n} row{'s' if n != 1 else ''}")
            self._log(f"Loaded {n} tasks from '{fname}'.", "info")
            self._update_stats()
            self.btn_run.setEnabled(True)
            self._rib_run.setEnabled(True)
            self.statusBar().showMessage(f"  Loaded: {fname}  —  {n} tasks")

        except Exception as ex:
            self._log(f"Failed to load Excel: {ex}", "error")
            QMessageBox.critical(self, "Load Error", str(ex))

    def _cell(self, text: str) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        item.setForeground(QColor(TEXT))
        return item

    # ── Automation control ────────────────────────────────────────────────────

    def _run_automation(self):
        if not self._tasks:
            QMessageBox.information(self, "No Tasks", "Load an Excel file first.")
            return

        for i in range(self.table.rowCount()):
            w = self.table.cellWidget(i, 4)
            if isinstance(w, StatusBadge):
                w.set_status("Pending")
            self.table.setItem(i, 5, self._cell(""))
        for r in self._results:
            r["status"] = "Pending"
            r["result"] = ""

        self.progress.setValue(0)
        self.btn_run.setEnabled(False)
        self._rib_run.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self._rib_stop.setEnabled(True)
        self.btn_export.setEnabled(False)
        self._rib_export.setEnabled(False)
        self.statusBar().showMessage("  Running automation…")
        self._log("─" * 48, "muted")
        self._log(f"Starting {len(self._tasks)} task(s)…", "info")

        headless = self.browser_mode.currentIndex() == 0
        self._worker = AutomationWorker(self._tasks, headless)
        self._thread = QThread()
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.log.connect(self._log)
        self._worker.row_update.connect(self._on_row_update)
        self._worker.progress.connect(self.progress.setValue)
        self._worker.finished.connect(self._on_finished)
        self._thread.start()

    def _stop_automation(self):
        if self._worker:
            self._worker.stop()
        self.btn_stop.setEnabled(False)
        self._rib_stop.setEnabled(False)
        self.statusBar().showMessage("  Stopping…")

    def _on_row_update(self, row_idx: int, status: str, result: str):
        if row_idx < self.table.rowCount():
            w = self.table.cellWidget(row_idx, 4)
            if isinstance(w, StatusBadge):
                w.set_status(status)
            self.table.setItem(row_idx, 5, self._cell(result))
        if row_idx < len(self._results):
            self._results[row_idx]["status"] = status
            self._results[row_idx]["result"] = result
        self._update_stats()

    def _on_finished(self, success: bool, msg: str):
        self.btn_run.setEnabled(True)
        self._rib_run.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self._rib_stop.setEnabled(False)
        self.btn_export.setEnabled(True)
        self._rib_export.setEnabled(True)
        level = "info" if success else "error"
        self._log(f"{'✔' if success else '✖'}  {msg}", level)
        self._log("─" * 48, "muted")
        self.statusBar().showMessage(f"  {msg}")
        self._update_stats()
        if self._thread:
            self._thread.quit()
            self._thread.wait()

    # ── Export ────────────────────────────────────────────────────────────────

    def _export_excel(self):
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "Missing Library",
                "openpyxl is not installed.\n\npip install openpyxl")
            return

        default_path = ""
        if self._excel_path:
            p = Path(self._excel_path)
            default_path = str(p.parent / f"{p.stem}_results{p.suffix}")

        save_path, _ = QFileDialog.getSaveFileName(
            self, "Export Results", default_path or "results.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not save_path:
            return

        try:
            if self._excel_path and Path(self._excel_path).exists():
                wb = openpyxl.load_workbook(self._excel_path)
            else:
                wb = openpyxl.Workbook()

            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(h).strip().lower().replace(" ", "_") if h else ""
                       for h in rows[0]] if rows else []

            def _col_for(name):
                for i, h in enumerate(headers):
                    if h == name:
                        return i + 1
                ws.cell(row=1, column=len(headers) + 1, value=name.capitalize())
                headers.append(name)
                return len(headers)

            status_col = _col_for("status")
            result_col = _col_for("result")
            ts_col     = _col_for("timestamp")

            ts        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            done_fill = PatternFill("solid", fgColor="E8F5E9")
            err_fill  = PatternFill("solid", fgColor="FFEBEE")
            pend_fill = PatternFill("solid", fgColor="FFFDE7")

            for i, r in enumerate(self._results):
                excel_row = i + 2
                status    = r.get("status", "Pending")
                result    = r.get("result", "")
                ws.cell(row=excel_row, column=status_col, value=status)
                ws.cell(row=excel_row, column=result_col, value=result)
                ws.cell(row=excel_row, column=ts_col,     value=ts)
                fill = done_fill if status == "Done" else (err_fill if status == "Error" else pend_fill)
                for c in [status_col, result_col, ts_col]:
                    ws.cell(row=excel_row, column=c).fill = fill

            wb.save(save_path)
            self._log(f"Exported to: {save_path}", "info")
            self.statusBar().showMessage(f"  Exported → {save_path}")
            QMessageBox.information(self, "Export Complete",
                f"Results saved to:\n{save_path}")

        except Exception as ex:
            self._log(f"Export failed: {ex}", "error")
            QMessageBox.critical(self, "Export Error", str(ex))

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _log(self, msg: str, level: str = "info"):
        colours = {
            "info":  "#58A6FF",
            "warn":  "#D29922",
            "error": "#F85149",
            "muted": "#8B949E",
        }
        c  = colours.get(level, "#E6EDF3")
        ts = datetime.now().strftime("%H:%M:%S")
        html = (
            f'<span style="color:#8B949E;">[{ts}]</span> '
            f'<span style="color:{c};">{msg}</span>'
        )
        self.log_box.append(html)
        self.log_box.verticalScrollBar().setValue(
            self.log_box.verticalScrollBar().maximum()
        )

    def _update_stats(self):
        total   = len(self._results)
        done    = sum(1 for r in self._results if r["status"] == "Done")
        errors  = sum(1 for r in self._results if r["status"] == "Error")
        pending = sum(1 for r in self._results if r["status"] == "Pending")
        self.stat_total.setText(str(total))
        self.stat_done.setText(str(done))
        self.stat_error.setText(str(errors))
        self.stat_pending.setText(str(pending))


# ══════════════════════════════════════════════════════════════════════════════
#  Entry point
# ══════════════════════════════════════════════════════════════════════════════

def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    pal = QPalette()
    pal.setColor(QPalette.Window,          QColor(BG))
    pal.setColor(QPalette.WindowText,      QColor(TEXT))
    pal.setColor(QPalette.Base,            QColor(SURFACE))
    pal.setColor(QPalette.AlternateBase,   QColor(BG))
    pal.setColor(QPalette.Text,            QColor(TEXT))
    pal.setColor(QPalette.Button,          QColor(SURFACE))
    pal.setColor(QPalette.ButtonText,      QColor(TEXT))
    pal.setColor(QPalette.Highlight,       QColor(BLUE))
    pal.setColor(QPalette.HighlightedText, QColor("#FFFFFF"))
    app.setPalette(pal)

    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()